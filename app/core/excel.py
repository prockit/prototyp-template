"""Excel import and export. Departments live in Excel; "load my list" is usually the first request.

Imported rows always go through a feature service (validated), never straight into a table.
Real personal data in an import is anonymised before it is stored.
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook


def read_rows(path: str | Path) -> list[dict[str, object]]:
    """Read the first sheet. The first row is the header; empty rows are skipped."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        header_cells = next(rows, ())
        header = [
            str(cell).strip() if cell is not None else f"column_{index + 1}"
            for index, cell in enumerate(header_cells)
        ]
        result: list[dict[str, object]] = []
        for row in rows:
            if all(cell is None for cell in row):
                continue
            result.append(dict(zip(header, row, strict=False)))
        return result
    finally:
        workbook.close()


def write_rows(
    path: str | Path, rows: list[dict[str, object]], sheet_title: str = "Export"
) -> Path:
    """Write dictionaries as a sheet; the keys of the first row form the header."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    if rows:
        header = list(rows[0].keys())
        sheet.append(header)
        for row in rows:
            sheet.append([row.get(column) for column in header])
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target
